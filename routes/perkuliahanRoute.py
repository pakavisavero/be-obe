from fastapi import Form, Request, BackgroundTasks, File, UploadFile
from fastapi import Depends, status, Header

from controller import perkuliahan, mataKuliah, tahunAjaran
from routes.route import app
from controller.utils import help_filter, check_access_module, remove_char

from db.session import db, getUsername
from controller.utils import decode_token
from db.database import Session
from db.schemas.perkuliahanSchema import (
    PerkuliahanResponseSchema,
    PerkuliahanDeleteSchema,
)

from HandlerCustom import HandlerCustom
from db.models import *

from openpyxl import load_workbook
from fastapi.responses import FileResponse
from starlette.background import BackgroundTasks
from datetime import datetime
from controller.utils import DocStatus
from openpyxl.utils import get_column_letter

import os
import shutil
import subprocess
import bcrypt


PERKULIAHAN = "/api/perkuliahan"
MODULE_NAME = "KBM Aktif"


def errArray(idx):
    if idx < 2:
        return 0
    else:
        return 1


@app.get(PERKULIAHAN + "s", response_model=PerkuliahanResponseSchema)
@check_access_module
async def get_all_perkuliahan(
    db: Session = Depends(db),
    token: str = Header(default=None),
    request: Request = None,
    page: int = 0,
    module_access=MODULE_NAME,
):
    filtered_data = help_filter(request)
    if filtered_data:
        query = perkuliahan.getAllPagingFiltered(
            db, page, filtered_data, token, xtra={'is_active': True})

        return {
            "code": status.HTTP_200_OK,
            "message": "Success retrieve filtered perkuliahan",
            "data": query["data"],
            "total": query["total"],
        }
    else:
        query = perkuliahan.getAllPaging(
            db, page, token, xtra={'is_active': True})
        return {
            "code": status.HTTP_200_OK,
            "message": "Success retrieve all perkuliahan",
            "data": query["data"],
            "total": query["total"],
        }


@app.get("/api/history-kbm", response_model=PerkuliahanResponseSchema)
@check_access_module
async def get_all_history_kbm(
    db: Session = Depends(db),
    token: str = Header(default=None),
    request: Request = None,
    page: int = 0,
    module_access=MODULE_NAME,
):
    filtered_data = help_filter(request)
    if filtered_data:
        query = perkuliahan.getAllPagingFiltered(
            db, page, filtered_data, token, xtra={'is_active': False}
        )

        return {
            "code": status.HTTP_200_OK,
            "message": "Success retrieve filtered history kbm",
            "data": query["data"],
            "total": query["total"],
        }
    else:
        query = perkuliahan.getAllPaging(
            db, page, token, xtra={'is_active': False})
        return {
            "code": status.HTTP_200_OK,
            "message": "Success retrieve all history kbm",
            "data": query["data"],
            "total": query["total"],
        }


@app.get(PERKULIAHAN + "/{id}", response_model=PerkuliahanResponseSchema)
@check_access_module
async def get_perkuliahan(
    db: Session = Depends(db),
    token: str = Header(default=None),
    id: int = None,
    request: Request = None,
    module_access=MODULE_NAME,
):
    data = perkuliahan.getByID(db, id, token)
    return {
        "code": status.HTTP_200_OK,
        "message": "Success get perkuliahan",
        "data": data,
    }


@app.post(PERKULIAHAN, response_model=PerkuliahanResponseSchema)
@check_access_module
async def submit_perkuliahan(
    db: Session = Depends(db),
    token: str = Header(default=None),
    data: dict = None,
    request: Request = None,
    module_access=MODULE_NAME,
):
    username = getUsername(token)

    res = perkuliahan.create(db, username, data)
    if res:
        return {
            "code": status.HTTP_200_OK,
            "message": "Success submit perkuliahan",
            "data": res,
        }

    else:
        return {
            "code": status.HTTP_400_BAD_REQUEST,
            "message": "error submit perkuliahan",
        }


@app.post('/api/deactivate-partial')
# @check_access_module
async def deactivate_perkuliahan_partial(
    db: Session = Depends(db),
    token: str = Header(default=None),
    data: dict = None,
    request: Request = None,
    module_access=MODULE_NAME,
):
    username = getUsername(token)
    db.query(Perkuliahan).filter_by(
        doc_status_id=3).update({'is_active': False})
    db.commit()

    return {
        "code": status.HTTP_200_OK,
        "message": "Success deactivate sebagian perkuliahan!",
    }


@app.post('/api/deactivate-all')
# @check_access_module
async def deactivate_perkuliahan_all(
    db: Session = Depends(db),
    token: str = Header(default=None),
    data: dict = None,
    request: Request = None,
    module_access=MODULE_NAME,
):
    username = getUsername(token)

    pkActive = db.query(Perkuliahan).filter_by(is_active=True).all()
    partial = False
    is_done = True

    for pk in pkActive:
        if pk.doc_status_id != 3:
            is_done = False
        else:
            partial = True

    if not is_done:
        return {
            "code": status.HTTP_400_BAD_REQUEST,
            "partial": partial,
            "message": "Terdapat perkuliahan yang masih dalam proses!",
        }

    else:
        db.query(Perkuliahan).update({'is_active': False})
        db.commit()
        return {
            "code": status.HTTP_200_OK,
            "message": "Success deactivate perkuliahan!",
        }


@app.put(PERKULIAHAN, response_model=PerkuliahanResponseSchema)
@check_access_module
async def update_perkuliahan(
    db: Session = Depends(db),
    token: str = Header(default=None),
    data: dict = None,
    request: Request = None,
    module_access=MODULE_NAME,
):
    username = getUsername(token)
    res = perkuliahan.update(db, username, data)
    if res:
        return {
            "code": status.HTTP_200_OK,
            "message": "Success update perkuliahan",
            "data": data,
        }
    else:
        return {
            "code": status.HTTP_400_BAD_REQUEST,
            "message": "error update perkuliahan",
        }


@app.delete(PERKULIAHAN)
async def delete_perkuliahan(
    db: Session = Depends(db),
    token: str = Header(default=None),
    data: PerkuliahanDeleteSchema = None,
):
    return {
        "code": status.HTTP_200_OK,
        "message": "Success delete perkuliahan",
    }


@app.get('/api/get-perkuliahan-by-matkul/{id}')
async def get_perkuliahan_by_matkul(
    db: Session = Depends(db),
    token: str = Header(default=None),
    request: Request = None,
    id: int = None,
):
    data = db.query(Perkuliahan).\
        filter_by(mata_kuliah_id=id).\
        filter_by(doc_status_id=3).\
        all()

    for dt in data:
        setattr(dt, 'mataKuliah', dt.mataKuliah)
        setattr(dt, 'tahunAjaran', dt.tahunAjaran)
        setattr(dt, 'prodi', dt.prodi)
        setattr(dt, 'docstatus', dt.docstatus)

        cpmk = db.query(CPMK).filter_by(perkuliahan_id=dt.id).all()
        matkulInfo = []

        for cp in cpmk:
            dictCPMK = {
                'name': cp.name,
                'statement': cp.statement,
                'cpl': []
            }

            mapping = db.query(MappingCpmkCpl).filter_by(cpmk_id=cp.id).all()
            for map in mapping:
                cpl = db.query(CPL).filter_by(id=map.cpl_id).first()
                dictCPMK['cpl'].append(cpl)

            matkulInfo.append(dictCPMK)

        setattr(dt, 'cpmk', matkulInfo)

    return {
        'status': status.HTTP_200_OK,
        'message': 'Success retrieve data!',
        'data': data
    }


@app.post(PERKULIAHAN + "-upload")
async def upload(
    db: Session = Depends(db),
    token: str = Header(default=None),
    data: dict = None,
    request: Request = None,
):
    try:
        username = getUsername(token)

        data = data["data"]
        thn_ajaran = str(data[0][1]).replace(":", "").strip()
        semester = str(data[1][1]).replace(":", "").strip()

        if semester == "" or thn_ajaran == "":
            raise ValueError(
                "Tidak terdapat semester atau tahun ajaran pada row " +
                str(idx)
            )

        headerLen = 10
        idx = 0

        for value in data:
            if idx >= 4 and len(value) > 0:
                dis = headerLen - len(value)
                if dis > 0:
                    for _ in range(0, dis):
                        value.append("")

                matkul = str(value[0]).strip()
                kode = str(value[1]).strip()
                kelas = str(value[2]).strip()
                sks = str(value[3])
                dosen1 = str(value[4]).strip()
                nip_dosen1 = str(value[5]).strip()
                dosen2 = value[6]
                nip_dosen2 = value[7]
                dosen3 = value[8]
                nip_dosen3 = value[9]

                if (
                    matkul == "" or
                    kode == "" or
                    nip_dosen1 == "" or
                    sks == '' or
                    matkul == 'None' or
                    kode == 'None' or
                    nip_dosen1 == 'None' or
                    sks == 'None'
                ):
                    idx += 1
                    break
                    # raise ValueError(
                    #     "Tidak terdapat mata kuliah, kode, SKS atau NIP dosen utama pada row "
                    #     + str(idx)
                    # )

                sks = int(sks.strip().lower().replace("sks", ""))

                isMatkulExist = db.query(
                    MataKuliah).filter_by(kode_mk=kode).first()
                if not isMatkulExist:
                    crMatkul = MataKuliah(
                        **{
                            "kode_mk": kode,
                            "mata_kuliah": matkul.lower().title(),
                            "sks": sks,
                            "kurikulum_id": 1,
                            "prodi_id": 8,
                        }
                    )
                    db.add(crMatkul)
                    db.commit()
                    db.refresh(crMatkul)

                    if crMatkul:
                        isMatkulExist = crMatkul
                    else:
                        raise ValueError(
                            "Tidak terdapat mata kuliah yang pada row " +
                            str(idx)
                        )

                ta = db.query(TahunAjaran).filter_by(
                    tahun_ajaran=thn_ajaran).first()
                if not ta:
                    head, _, _ = thn_ajaran.partition("/")
                    ta = TahunAjaran(
                        **{"name": head, "tahun_ajaran": thn_ajaran, "is_active": True}
                    )
                    db.add(ta)
                    db.commit()
                    db.refresh(ta)

                data = {
                    "prodi_id": 8,
                    "mata_kuliah_id": isMatkulExist.id,
                    "kelas": kelas,
                    "tahun_ajaran_id": ta.id,
                    "semester": semester,
                    "is_active": True,
                    "created_by": "Import User",
                    "modified_by": "Import User",
                }

                existDosen1 = db.query(User).filter_by(nip=nip_dosen1).first()
                if not existDosen1:
                    salt = bcrypt.gensalt()
                    bytes = "test@2023".encode("utf-8")
                    password = bcrypt.hashpw(bytes, salt)

                    existDosen1 = User(
                        **{
                            "email": dosen1,
                            "nip": nip_dosen1,
                            "full_name": dosen1,
                            "prodi_id": 8,
                            "password": password,
                            "created_by": username,
                            "modified_by": username,
                        }
                    )
                    db.add(existDosen1)
                    db.commit()
                    db.refresh(existDosen1)

                    crUserRole = UserRole(
                        **{
                            "user_id": existDosen1.id,
                            "role_id": 3,
                            "created_by": username,
                            "modified_by": username,
                        }
                    )
                    db.add(crUserRole)
                    db.commit()
                    db.refresh(crUserRole)

                data["dosen_id"] = existDosen1.id
                data["pj_dosen_id"] = existDosen1.id
                data["doc_status_id"] = int(
                    DocStatus.MENUNGGU_UPLOAD_DPNA.value)

                if nip_dosen2:
                    existDosen2 = db.query(User).filter_by(
                        nip=nip_dosen2).first()
                    if existDosen2:
                        data["dosen2_id"] = existDosen2.id

                if nip_dosen3:
                    existDosen3 = db.query(User).filter_by(
                        nip=nip_dosen3).first()
                    if existDosen3:
                        data["dosen3_id"] = existDosen3.id

                isExistPK = (
                    db.query(Perkuliahan)
                    .filter_by(mata_kuliah_id=data["mata_kuliah_id"])
                    .filter_by(tahun_ajaran_id=ta.id)
                    .filter_by(semester=data["semester"])
                    .filter_by(kelas=data["kelas"])
                    .first()
                )

                if not isExistPK:
                    pk = Perkuliahan(**data)
                    db.add(pk)
                    db.commit()
                    db.refresh(pk)

            idx += 1

        return {
            "code": status.HTTP_200_OK,
            "message": "Success Upload Perkuliahan",
        }

    except Exception as e:
        print(e)
        db.rollback()
        err = str(e.args[0]).split("\n")
        data = {"message": err[errArray(len(err))]}
        return {
            "code": status.HTTP_400_BAD_REQUEST,
            "message": err,
        }


@app.post("/api/dpna" + "-upload")
async def upload_dpna(
    db: Session = Depends(db),
    token: str = Header(default=None),
    data: dict = None,
    request: Request = None,
):
    user = decode_token(token)["fullName"]
    id = data["id"]
    header = data["data"]

    try:
        c1 = header[0][1]
        if len(c1.split("-")) != 2:
            raise ValueError("Format cell B1 (Kode MK - Matkul) tidak valid!")

        kode_mk = str(c1.split("-")[0]).replace(":", "").strip()
        tahun_ajaran = str(header[1][1]).replace(":", "").strip()
        semester = str(header[2][1]).replace(":", "").strip()
        kelas = str(header[3][1]).replace(":", "").strip()

        checkMatkul = db.query(MataKuliah).filter_by(kode_mk=kode_mk).first()
        if not checkMatkul:
            raise ValueError("Mata Kuliah tidak tersedia!")

        ta = db.query(TahunAjaran).filter_by(name=tahun_ajaran).first()
        if not ta:
            raise ValueError("Tahun Ajaran tidak tersedia!")

        pkActive = db.query(Perkuliahan).filter_by(id=id).first()
        if (
            (pkActive.mata_kuliah_id != checkMatkul.id)
            or (pkActive.tahun_ajaran_id != ta.id)
            or (pkActive.semester != semester)
            or (pkActive.kelas != kelas)
        ):
            raise ValueError("Data Perkuliahan tidak selaras!")

        checkPk = (
            db.query(Perkuliahan)
            .filter_by(mata_kuliah_id=checkMatkul.id)
            .filter_by(tahun_ajaran_id=ta.id)
            .filter_by(semester=semester)
            .filter_by(kelas=kelas)
            .first()
        )

        setattr(checkPk, "doc_status_id", int(
            DocStatus.MENUNGGU_UPLOAD_CPMK.value))
        db.commit()

        if not checkPk:
            raise ValueError("Perkuliahan tidak tersedia!")

        idx = 0
        headerLen = 5
        for value in data["data"]:
            if idx >= 7:
                dis = headerLen - len(value)
                if dis > 0:
                    for _ in range(0, dis):
                        value.append("")

                nim = str(value[0])
                full_name = str(value[1])

                if nim == '' and full_name == '':
                    idx += 1
                    break

                semMhs = int(value[2])
                statusMhs = str(value[3])

                is_valid = True
                if "75" in full_name:
                    is_valid = False

                char = ["(", "<", "%", ">", ")"]
                for ch in char:
                    full_name = full_name.replace(ch, "")

                if nim and full_name:
                    mhs = db.query(Mahasiswa).filter_by(nim=nim).first()
                    if not mhs:
                        mhs = Mahasiswa(
                            **{
                                "nim": nim,
                                "full_name": full_name.rsplit(" ", 1)[0]
                                .lower()
                                .title()
                                .strip(),
                                "prodi_id": 8,
                                "status_mhs_id": 1,
                                "semester": semMhs,
                                "created_by": user,
                                "modified_by": user,
                            }
                        )
                        db.add(mhs)
                        db.commit()
                        db.refresh(mhs)

                    db.query(Mahasiswa).update({"semester": semMhs})
                    mapping = MappingMahasiswa(
                        **{
                            "perkuliahan_id": checkPk.id,
                            "mahasiswa_id": mhs.id,
                            "status": statusMhs,
                            "is_valid": is_valid,
                            "created_by": user,
                            "modified_by": user,
                        }
                    )
                    db.add(mapping)
                    db.commit()

            idx += 1

        persenTugas = remove_char(header[6][4])
        persenPraktek = remove_char(header[6][5])
        persenUts = remove_char(header[6][6])
        persenUas = remove_char(header[6][7])

        presentase = PresentasePK(
            **{
                "perkuliahan_id": id,
                "nilai_tugas": persenTugas,
                "nilai_uts": persenUts,
                "nilai_uas": persenUas,
                "nilai_praktek": persenPraktek,
                "is_active": True,
            }
        )
        db.add(presentase)
        db.commit()

        return {
            "code": status.HTTP_200_OK,
            "message": "Success Upload DPNA",
        }

    except Exception as e:
        print(e)
        db.rollback()
        err = str(e.args[0]).split("\n")
        data = {"message": err[errArray(len(err))]}
        raise HandlerCustom(data=data)


@app.get("/api/get-template/{id}", response_model=PerkuliahanResponseSchema)
async def bg_task_template(
    db: Session = Depends(db),
    background_tasks=BackgroundTasks,
    token: str = Header(default=None),
    id: int = None,
):
    checkExpDPNA = db.query(CheckExportDPNA).filter_by(
        perkuliahan_id=id).first()
    if not checkExpDPNA:
        background_tasks.add_task(get_template(db, id))
    else:
        dir = checkExpDPNA.template_name
        return FileResponse(path=dir, filename=dir.replace("files/", ""))


@app.post('/api/save-template')
async def save_template(
    db: Session = Depends(db),
    token: str = Header(default=None),
    file: UploadFile = File(...),
    id: int = Form()
):
    try:
        contents = file.file.read()
        with open(file.filename, 'wb') as f:
            f.write(contents)

        path = "files/cpmk/{}".format(file.filename)
        shutil.move(file.filename, path)

        wb = load_workbook(path, data_only=True)
        sheets = [
            'Do',
            'RPS-old',
            'PETUNJUK',
            'CPMK-CPL',
            'NILAI TUGAS',
            'NILAI PRAKTEK',
            'NILAI UTS',
            'NILAI UAS',
            'FORM NILAI SIAP',
            'CPL',
            'CPMK',
            'Evaluasi',
        ]

        for sheet in sheets:
            try:
                del wb[sheet]
            except:
                pass

        RPS = wb["RPS"]
        statKontrak = ''
        KONTRAK = ''
        try:
            statKontrak = "KONTRAK"
            KONTRAK = wb["KONTRAK"]
        except Exception:
            statKontrak = "Kontrak"
            KONTRAK = wb["Kontrak"]

        pk = db.query(Perkuliahan).filter_by(id=id).first()
        dosen1 = ''
        dosen2 = ''
        dosen3 = ''
        tableCpmk = []

        cell = 29
        for row in range(cell, 42):
            tableCpmk.append([
                KONTRAK['D' + str(row)].value,
                KONTRAK['E' + str(row)].value,
                KONTRAK['F' + str(row)].value,
                KONTRAK['G' + str(row)].value,
            ])

        if pk.dosen1:
            dosen1 = pk.dosen1.full_name
        if pk.dosen2:
            dosen2 = pk.dosen2.full_name
        if pk.dosen3:
            dosen3 = pk.dosen3.full_name

        cplValue = ''
        cpmkValue = ''
        idsCpl = []

        cpmks = db.query(CPMK).filter_by(perkuliahan_id=id).all()
        for cpmk in cpmks:
            cpmkValue += cpmk.name + '   ' + cpmk.statement + '\n'
            mapping = db.query(MappingCpmkCpl).filter_by(cpmk_id=cpmk.id).all()
            for map in mapping:
                if map.cpl_id not in idsCpl:
                    idsCpl.append(map.cpl_id)
                    cplValue += '[ ' + map.cpl.name + ' ]' + \
                        map.cpl.statement + '\n'

        RPS["E5"] = pk.mataKuliah.kode_mk + ' - ' + pk.mataKuliah.mata_kuliah
        RPS["L5"] = str(pk.mataKuliah.sks) + ' ' + 'SKS'
        RPS["O5"] = pk.semester
        RPS["E7"] = dosen1 + ' / ' + dosen2 + ' / ' + dosen3
        RPS["E8"] = cplValue
        RPS["E9"] = cpmkValue

        KONTRAK["E4"] = pk.mataKuliah.kode_mk + \
            ' - ' + pk.mataKuliah.mata_kuliah
        KONTRAK["E5"] = pk.tahunAjaran.tahun_ajaran
        KONTRAK["E6"] = dosen1 + ' / ' + dosen2 + ' / ' + dosen3
        KONTRAK["H5"] = pk.semester
        KONTRAK["A16"] = cplValue
        KONTRAK["A18"] = cpmkValue

        row = 29
        for table in tableCpmk:
            KONTRAK['D' + str(row)] = table[0]
            KONTRAK['E' + str(row)] = table[1]
            KONTRAK['F' + str(row)] = table[2]
            KONTRAK['G' + str(row)] = table[3]
            row += 1

        for sheet in [statKontrak, 'RPS', 'COVER']:
            for idx, col in enumerate(wb[sheet].columns, 1):
                wb[sheet].column_dimensions[get_column_letter(
                    idx)].auto_size = True

        wb['COVER'].row_dimensions[8].height = 25
        wb['COVER'].row_dimensions[9].height = 25
        wb['RPS'].row_dimensions[9].height = 100
        KONTRAK.row_dimensions[13].height = 70
        KONTRAK.row_dimensions[20].height = 250

        wb.save(path)

        pdfName = file.filename.replace('xlsx', 'pdf')
        exp = CheckExportPortofolio(
            **{
                "perkuliahan_id": id,
                "template_name": pdfName,
            }
        )
        db.add(exp)
        db.commit()

        subprocess.run(["libreoffice", "--headless",
                        "macro:///Standard.Module1.FitToPage", "--convert-to", "pdf", path])
        shutil.move(pdfName, 'files/cpmk/' + pdfName)

        remove_file(path)

    except Exception:
        return {"message": "There was an error uploading the file"}

    finally:
        file.file.close()

    return {"message": f"Successfully uploaded {file.filename}"}


@app.get("/api/get-portofolio" + "/{id}")
async def get_portofolio(
    db: Session = Depends(db),
    token: str = Header(default=None),
    request: Request = None,
    id: int = None
):
    portName = db.query(CheckExportPortofolio).filter_by(
        perkuliahan_id=id).first()

    path = perkuliahan.getJinjaPortofolio(
        db, request, id, portName.template_name)
    headers = {'Content-Disposition': 'inline; filename="out.pdf"'}

    return FileResponse(
        path,
        headers=headers,
        media_type='application/pdf'
    )


def get_template(db: Session, id: int):
    pk = db.query(Perkuliahan).filter_by(id=id).first()
    matkul = mataKuliah.getByID(db, pk.mata_kuliah_id, "token")
    ta = tahunAjaran.getByID(db, pk.tahun_ajaran_id, "token")
    mapping = db.query(MappingMahasiswa).filter_by(perkuliahan_id=pk.id).all()

    wb = load_workbook("files/template_cpmk.xlsx")

    SH_CPMK = wb["CPMK-CPL"]
    SH_TUGAS = wb["NILAI TUGAS"]
    SH_PRAKTEK = wb["NILAI PRAKTEK"]
    SH_UTS = wb["NILAI UTS"]
    SH_UAS = wb["NILAI UAS"]
    SH_SIAP = wb["FORM NILAI SIAP"]

    SH_CPMK["C1"] = matkul.kode_mk
    SH_CPMK["C2"] = ta.tahun_ajaran
    SH_CPMK["C3"] = pk.semester
    SH_CPMK["C4"] = pk.kelas

    no = 1
    row = 8
    for map in mapping:
        mhs = db.query(Mahasiswa).filter_by(id=map.mahasiswa_id).first()
        full_name = mhs.full_name.upper()
        if not map.is_valid:
            full_name += " (< 75%)"

        SH_TUGAS["A" + str(row)] = no
        SH_TUGAS["B" + str(row)] = mhs.nim.upper()
        SH_TUGAS["C" + str(row)] = full_name

        SH_PRAKTEK["A" + str(row)] = no
        SH_PRAKTEK["B" + str(row)] = mhs.nim
        SH_PRAKTEK["C" + str(row)] = full_name

        SH_UTS["A" + str(row)] = no
        SH_UTS["B" + str(row)] = mhs.nim
        SH_UTS["C" + str(row)] = full_name

        SH_UAS["A" + str(row)] = no
        SH_UAS["B" + str(row)] = mhs.nim
        SH_UAS["C" + str(row)] = full_name

        SH_SIAP["A" + str(row)] = mhs.nim
        SH_SIAP["B" + str(row)] = full_name

        no += 1
        row += 1

    for col in ["E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P", "Q", "R"]:
        SH_TUGAS.column_dimensions[col].hidden = True
        SH_PRAKTEK.column_dimensions[col].hidden = True
        SH_UTS.column_dimensions[col].hidden = True
        SH_UAS.column_dimensions[col].hidden = True

    now = datetime.now().strftime("%Y_%m_%d-%I_%M_%S")
    path = "files/{}_{}_{}.xlsx".format(
        (matkul.mata_kuliah).lower().replace(
            " ", "_"), pk.semester.lower(), now
    )
    wb.save(path)

    exp = CheckExportDPNA(
        **{
            "perkuliahan_id": id,
            "template_name": path,
        }
    )
    db.add(exp)
    db.commit()


@app.post("/api/cpmk" + "-upload")
async def upload_cpmk(
    db: Session = Depends(db),
    token: str = Header(default=None),
    request: Request = None,
    data: dict = None,
):
    id = data["id"]
    data = data["data"]

    try:
        SH_CPMK = data[0]
        SH_TUGAS = data[1]
        SH_PRAKTEK = data[2]
        SH_UTS = data[3]
        SH_UAS = data[4]
        SH_SIAP = data[5]
        SH_CPMK_MHS = data[6]
        SH_CPL_MHS = data[7]
        SH_EVALUASI = data[8]

        # CPMK
        c1 = SH_CPMK[0][2]
        if len(c1.split("-")) != 2:
            raise ValueError("Format cell C1 (Kode MK - Matkul) tidak valid!")

        kode_mk = str(c1.split("-")[0]).replace(":", "").strip()
        tahun_ajaran = str(SH_CPMK[1][2]).replace(":", "").strip()
        semester = str(SH_CPMK[2][2]).replace(":", "").strip()
        kelas = str(SH_CPMK[3][2]).replace(":", "").strip()
        sks = str(SH_CPMK[8][2]).replace(":", "").strip()

        # Check If Related
        checkMatkul = db.query(MataKuliah).filter_by(kode_mk=kode_mk).first()
        if not checkMatkul:
            raise ValueError("Mata Kuliah tidak tersedia!")

        ta = db.query(TahunAjaran).filter_by(name=tahun_ajaran).first()
        if not ta:
            raise ValueError("Tahun Ajaran tidak tersedia!")

        pkActive = db.query(Perkuliahan).filter_by(id=id).first()
        if (
            (pkActive.mata_kuliah_id != checkMatkul.id)
            or (pkActive.tahun_ajaran_id != ta.id)
            or (pkActive.semester != semester)
            or (pkActive.kelas != kelas)
        ):
            raise ValueError("Data Perkuliahan tidak selaras!")

        checkPk = (
            db.query(Perkuliahan)
            .filter_by(mata_kuliah_id=checkMatkul.id)
            .filter_by(tahun_ajaran_id=ta.id)
            .filter_by(semester=semester)
            .filter_by(kelas=kelas)
            .first()
        )

        if not checkPk:
            raise ValueError("Perkuliahan tidak tersedia!")

        setattr(checkPk, "doc_status_id", int(DocStatus.SELESAI.value))
        db.commit()

        perkuliahan.insertCpl(db, token, checkPk.id, SH_CPMK)
        perkuliahan.insertCpmk(db, token, checkPk.id, SH_CPMK)
        perkuliahan.insertNilai(db, token, checkPk.id,
                                SH_TUGAS, NilaiTugas, "tugas")
        perkuliahan.insertNilai(
            db, token, checkPk.id, SH_PRAKTEK, NilaiPraktek, "praktek"
        )
        perkuliahan.insertNilai(db, token, checkPk.id, SH_UTS, NilaiUTS, "uts")
        perkuliahan.insertNilai(db, token, checkPk.id, SH_UAS, NilaiUAS, "uas")

        perkuliahan.insertCPLMahasiswa(db, token, checkPk.id, SH_CPL_MHS)
        perkuliahan.insertCPMKMahasiswa(db, token, checkPk.id, SH_SIAP)
        perkuliahan.insertEvaluasi(db, token, checkPk.id, SH_EVALUASI)

        return {
            "code": status.HTTP_200_OK,
            "message": "Success Upload CPMK",
        }

    except Exception as e:
        print(e)
        db.rollback()
        err = str(e.args[0]).split("\n")
        data = {"message": err[errArray(len(err))]}
        raise HandlerCustom(data=data)


def remove_file(path: str) -> None:
    os.unlink(path)


@app.get("/api/get-form-siap/{id}")
async def upload_form_siap(
    db: Session = Depends(db),
    token: str = Header(default=None),
    background_tasks: BackgroundTasks = None,
    id: int = None,
):
    path = perkuliahan.getNilaiSiap(db, id)
    background_tasks.add_task(remove_file, path)
    return FileResponse(path=path, filename=path.replace("files/siap/", ""))
